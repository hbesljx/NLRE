from modelscope import AutoModelForCausalLM, AutoTokenizer
import torch
from openpyxl import load_workbook

# 设备配置
device = "cuda" if torch.cuda.is_available() else "cpu"

# 模型缓存路径
model_cache_path = "/root/.cache/modelscope/hub/qwen/Qwen2.5-7B-Instruct"

# 输入和输出文件路径
input_data = './data/test_data.xlsx'
output_data = './data/test_data_output.xlsx'

# 加载模型和分词器
model = AutoModelForCausalLM.from_pretrained(
    model_cache_path,
    torch_dtype=torch.float16,  # 使用半精度浮点数以节省显存
    device_map="auto"           # 自动分配模型到多个 GPU
)
tokenizer = AutoTokenizer.from_pretrained(model_cache_path)

# 预热对话（固定上下文）
initial_messages = [
    {"role": "system", "content": "你是千问，由阿里云创造，你是一个有用的助手"},
    {"role": "user", "content": "匹配电子邮件地址"},
    {"role": "assistant", "content": "[a-zA-Z0-9.-]+@[a-zA-Z0-9]+(-[a-zA-Z0-9]+)*"},
    {"role": "user", "content": "匹配2位数字"},
    {"role": "assistant", "content": "\\d{2}"},
    {"role": "user", "content": "我输入自然语言描述，请你直接输出对应的正则表达式，不要任何解释或额外内容。"},
    {"role": "assistant", "content": ""}
]

def generate_regex(input_text):
    """
    根据输入文本生成正则表达式
    :param input_text: 输入文本
    :return: 生成的正则表达式
    """
    # 构建完整的对话消息
    messages = initial_messages + [{"role": "user", "content": input_text}]
    
    # 构建输入文本
    text = tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True,
    )
    
    # 将输入转换为模型可接受的格式
    model_inputs = tokenizer(text, return_tensors="pt", padding=True, truncation=True).to(device)
    
    # 生成输出
    generated_ids = model.generate(
        **model_inputs,
        max_new_tokens=50,      # 限制生成长度，避免过长输出
        temperature=0.1,        # 降低温度以提高确定性
        top_k=10,               # 设置 top-k 抽样
        top_p=0.9,              # 设置 nucleus 抽样
        use_cache=True          # 使用缓存以提高效率
    )
    
    # 解码生成的文本
    response = tokenizer.decode(
        generated_ids[0][len(model_inputs.input_ids[0]):],  # 只解码生成的部分
        skip_special_tokens=True
    ).strip()
    
    return response

def contains_chinese(text):
    """
    检查字符串是否包含中文字符
    :param text: 输入字符串
    :return: 是否包含中文字符
    """
    for char in text:
        if '\u4e00' <= char <= '\u9fff':
            return True
    return False

def main():
    # 加载 Excel 文件
    workbook = load_workbook(filename=input_data)
    sheet = workbook.active  # 获取活动工作表

    # 遍历每一行数据（跳过标题行）
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # 从第2行开始
        description, correct_regex = row

        # 如果目标列已经有内容，则跳过
        if sheet[f'C{row_idx}'].value:
            print(f"Skipping line {row_idx} (already processed)")
            continue

        # 如果有空行则跳过
        if not description or not correct_regex:
            print(f"Skipping empty row at line {row_idx}")
            continue

        # 调用模型生成正则表达式
        generated_regex = generate_regex(description)

        # 检查输出是否为空或包含中文字符
        if not generated_regex:
            print(f"Line {row_idx}: Empty output detected. Input: {description}")
        elif contains_chinese(generated_regex):
            print(f"Line {row_idx}: Chinese characters detected in output. Input: {description}, Output: {generated_regex}")

        # 打印生成结果
        print(f"Line {row_idx}: Input: {description}, Generated Regex: {generated_regex}")

        # 将生成的正则表达式写入第三列
        sheet[f'C{row_idx}'] = generated_regex

    # 保存修改后的 Excel 文件
    workbook.save(filename=output_data)
    print("Processing completed and saved to:", output_data)

if __name__ == "__main__":
    main()