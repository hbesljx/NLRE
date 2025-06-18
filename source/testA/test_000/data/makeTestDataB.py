import random
import string
from openpyxl import Workbook
import jieba

# 自定义同义词字典（扩充版）
custom_synonyms = {
    # 数字相关
    "数字": ["数值", "数位", "数目"],
    "匹配": ["符合", "对应", "适配", "满足", "达成"],
    "至少": ["最少", "不低于", "不少于"],
    "不超过": ["最多", "不多于", "小于等于"],
    "长度": ["大小", "尺寸", "字符数", "位数"],

    # 字母相关
    "小写": ["小写字母", "非大写", "lowercase"],
    "大写": ["大写字母", "非小写", "uppercase"],

    # 位置相关
    "开头": ["起始", "开始", "头部"],
    "结尾": ["结束", "末尾", "尾部"],
    "包含": ["包括", "涵盖", "涉及", "含有"],

    # 字符串结构相关
    "字符串": ["文本", "句子", "内容", "片段"],
    "任意": ["所有", "任何", "随机"],
    "恰好": ["刚好", "正好"],
    "固定": ["确定", "指定"],

    # 单词相关
    "组成": ["构成", "形成", "组合"],
    "由...组成": ["包含", "由...构成", "由...形成"],

    # 其他常用词汇
    "以...开头": ["起始于", "前缀为", "头部为"],
    "以...结尾": ["终止于", "后缀为", "尾部为"],
    "总长度": ["全长", "总体长度", "完整长度"],
    "前缀": ["开头部分", "前置部分", "头部内容"],
    "后缀": ["结尾部分", "后置部分", "尾部内容"],
}

# 定义基础模式，其中包含可以修改的部分
patterns = [
    {"description": "匹配{min}-{max}位数字", "regex": "^\\d{{{min},{max}}}$"},
    {"description": "匹配任意长度的数字", "regex": "\\d+"},
    {"description": "匹配至少{n}位数字", "regex": "^\\d{{{n},}}$"},  # 使用变量 {n}
    {"description": "匹配不超过{n}位数字", "regex": "^\\d{{1,{n}}}$"},  # 使用变量 {n}
    {"description": "匹配至少{min}位最多{max}位的小写字母", "regex": "^[a-z]{{{min},{max}}}$"},
    {"description": "匹配至少{min}位最多{max}位的大写字母", "regex": "^[A-Z]{{{min},{max}}}$"},
    {"description": "匹配{min}-{max}位的字母或数字", "regex": "^[a-zA-Z0-9]{{{min},{max}}}$"},
    {"description": "匹配以{prefix}开头且包含{word}的字符串", "regex": "^{prefix}.*{word}.*$"},
    # {"description": "匹配以{suffix}结尾且长度为{min}-{max}的字符串", "regex": "^.{{{min},{max}}}{suffix}$"},
    {"description": "匹配以{suffix}结尾且总长度为{total_min}-{total_max}的字符串","regex": "^.{{{prefix_min},{prefix_max}}}{suffix}$"},
    # {"description": "匹配由{min}-{max}个单词组成的句子，每个单词长度为{word_min}-{word_max}","regex": "^([a-zA-Z]{{{word_min},{word_max}}} ){{{min_minus_one},{max_minus_one}}}[a-zA-Z]{{{word_min},{word_max}}}$"},
    {"description": "匹配由{min}-{max}个单词组成的句子，每个单词长度为{word_min}-{word_max}","regex": "^([a-zA-Z]{{{word_min},{word_max}}}[ ,.!?;:'\"…]*){{{min_minus_one},{max_minus_one}}}[a-zA-Z]{{{word_min},{word_max}}}[.!?;:'\"…]?$"},
    {"description": "匹配含有{word1}和{word2}的字符串", "regex": "^(?=.*{word1})(?=.*{word2}).*$"},
    {"description": "匹配以{prefix}开头的任意长度字符串", "regex": "^{prefix}.*"},
    {"description": "匹配以{suffix}结尾的任意长度字符串", "regex": ".*{suffix}$"},
    # {"description": "匹配由{n}个单词组成的句子", "regex": "^([a-zA-Z]+\\s?){{{n}}}$"},  # 使用变量 {n}
    {"description": "匹配由{n}个单词组成的句子","regex": "^([a-zA-Z]+[ ,.!?;:'\"…]*){{{n_minus_one}}}[a-zA-Z]+[.!?;:'\"…]?$"}
]

# 新增模式：涵盖正则表达式的各种形式
new_patterns = [
    {"description": "匹配恰好{n}位数字", "regex": "^\\d{{{n}}}$"},  # 使用 \d 和 {}
    {"description": "匹配至少一个单词字符", "regex": "\\w+"},  # 使用 \w
    {"description": "匹配至少一个空白字符", "regex": "\\s+"},  # 使用 \s
    {"description": "匹配以数字开头的字符串", "regex": "^\\d.*$"},  # 使用 ^ 和 \d
    {"description": "匹配以字母结尾的字符串", "regex": ".*[a-zA-Z]$"},  # 使用 $ 和 [a-zA-Z]
    {"description": "匹配包含至少一个特殊字符的字符串", "regex": ".*[!@#$%^&*()].*"},  # 使用 []
    {"description": "匹配长度为{n}的单词", "regex": "^[a-zA-Z]{{{n}}}$"},  # 使用 {} 和 [a-zA-Z]
    {"description": "匹配不包含数字的字符串", "regex": "^[^\\d]*$"},  # 使用 [^...] 和 *
    {"description": "匹配以大写字母开头的单词", "regex": "^[A-Z][a-zA-Z]*$"},  # 使用 [A-Z] 和 [a-zA-Z]
    {"description": "匹配至少一个数字和一个字母的字符串", "regex": "^(?=.*\\d)(?=.*[a-zA-Z]).+$"},  # 使用 (?=...) 和 +
]

# 将新增模式合并到现有模式中
patterns.extend(new_patterns)

# 固定模式 - 包括中国大陆居民身份证号码、域名、Email地址等
fixed_patterns = [
    {"description": "匹配中国大陆居民身份证号码", "regex": "^\\d{17}[0-9Xx]$"},
    # {"description": "匹配域名", "regex": "[a-zA-Z0-9][-a-zA-Z0-9]{0,62}(\\.[a-zA-Z0-9][-a-zA-Z0-9]{0,62})+\\.?$"},
    # {"description": "匹配Email地址", "regex": "^\\w+([-+.]\\w+)*@\\w+([-.]\\w+)*\\.\\w+([-.]\\w+)*$"},
    # {"description": "匹配正整数", "regex": "^[1-9]\\d*$"},
    # {"description": "匹配负整数", "regex": "^-[1-9]\\d*$"},
    {"description": "匹配非负浮点数(允许有整数)", "regex": "^\\d+(\\.\\d+)?$"},
    {"description": "匹配非负浮点数(不允许有整数)", "regex": "^\\d+\\.\\d+$"},
    {"description": "匹配非正浮点数(包括负数和零)", "regex": "^(-?\\d+(\\.\\d+)?)|(0+(\\.0+)?)$"}
    # {"description": "匹配至少n位数字", "regex": "^\\d{n,}$"},
    # {"description": "匹配不超过n位数字", "regex": "^\\d{1,n}$"}
]


def generate_random_unicode_string(length):
    """生成随机的 Unicode 字符串（如中文字符）"""
    return ''.join(chr(random.randint(0x4E00, 0x9FFF)) for _ in range(length))


def add_noise_to_description(description):
    """在描述中加入随机噪声"""
    noise_phrases = ["例如", "比如", "可能包括", "通常用于"]
    if random.random() < 0.5:  # 50% 概率添加噪声
        noise = random.choice(noise_phrases)
        return f"{description} {noise}"
    return description


def synonym_replacement(text):
    """对描述中的关键字进行同义词替换（基于自定义字典）"""
    words = jieba.lcut(text)
    new_words = []
    for word in words:
        if word in custom_synonyms and random.random() < 0.5:  # 50% 概率替换
            new_word = random.choice(custom_synonyms[word])
            new_words.append(new_word)
        else:
            new_words.append(word)
    return "".join(new_words)


def replace_placeholders(pattern_info):
    """通用函数：替换描述和正则表达式中的占位符"""
    description = pattern_info['description']
    regex = pattern_info['regex']

    # 占位符映射表
    placeholders = {}

    # 替换 {prefix}
    if "{prefix}" in description:
        placeholders['prefix'] = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))

    # 替换 {suffix}
    if "{suffix}" in description:
        placeholders['suffix'] = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))

    # 替换 {word}
    if "{word}" in description:
        placeholders['word'] = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))

    # 替换 {min} 和 {max}
    if "{min}" in description or "{max}" in description:
        min_val = random.randint(1, 5)  # 最小单词数
        max_val = random.randint(min_val + 1, 10)  # 最大单词数
        placeholders['min'] = min_val
        placeholders['max'] = max_val

    # 计算 {min_minus_one} 和 {max_minus_one}
    placeholders['min_minus_one'] = max(0, placeholders.get('min', 1) - 1)
    placeholders['max_minus_one'] = max(0, placeholders.get('max', 1) - 1)

    # 替换 {word_min} 和 {word_max}
    if "{word_min}" in description or "{word_max}" in description:
        word_min = random.randint(1, 5)  # 每个单词的最小长度
        word_max = random.randint(word_min + 1, 10)  # 每个单词的最大长度
        placeholders['word_min'] = word_min
        placeholders['word_max'] = word_max

    # 替换 {word1} 和 {word2}
    if "{word1}" in description or "{word2}" in description:
        placeholders['word1'] = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))
        placeholders['word2'] = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))

    # 替换 {n}
    if "{n}" in description or "{n_minus_one}" in regex:
        n = random.randint(1, 10)  # 随机生成单词数量
        placeholders['n'] = n
        placeholders['n_minus_one'] = max(0, n - 1)

    # 替换 {total_min} 和 {total_max}
    if "{total_min}" in description or "{total_max}" in description:
        total_min = random.randint(7, 10)  # 总长度最小值
        total_max = random.randint(total_min + 1, 17)  # 总长度最大值
        placeholders['total_min'] = total_min
        placeholders['total_max'] = total_max

    # 替换 {suffix}
    if "{suffix}" in description:
        suffix = ''.join(random.choices(string.ascii_lowercase, k=random.randint(3, 8)))
        placeholders['suffix'] = suffix

    # 动态计算 {prefix_min} 和 {prefix_max}
    if "{prefix_min}" in regex or "{prefix_max}" in regex:
        suffix_length = len(placeholders.get('suffix', ''))
        prefix_min = max(0, placeholders.get('total_min', 7) - suffix_length)
        prefix_max = max(0, placeholders.get('total_max', 17) - suffix_length)
        placeholders['prefix_min'] = prefix_min
        placeholders['prefix_max'] = prefix_max

    # 替换所有占位符
    try:
        description = description.format(**placeholders)
        regex = regex.format(**placeholders)
    except KeyError as e:
        print(f"Error processing pattern: {pattern_info}. Missing key: {e}")
        return None

    return (description, regex)


def generate_unique_pattern(pattern):
    """生成唯一的正则表达式模式"""
    result = replace_placeholders(pattern)
    if result is None:
        return None

    description, regex = result

    # 数据增强：对描述进行同义词替换
    description = synonym_replacement(description)

    # 添加噪声
    description = add_noise_to_description(description)

    return (description, regex)


def generate_dataset(size=2700, max_attempts=40000):  # 增加尝试次数
    results = set()
    attempts = 0

    # 添加固定的模式
    for fixed_pattern in fixed_patterns:
        results.add((fixed_pattern["description"], fixed_pattern["regex"]))

    # 确保随机生成的模式填满剩余空间
    while len(results) < size and attempts < max_attempts:
        pattern_info = random.choice(patterns)
        entry = generate_unique_pattern(pattern_info)
        if entry is not None:  # 确保生成的 entry 不为 None
            results.add(entry)
        attempts += 1
    if attempts >= max_attempts:
        print("Warning: Reached maximum number of attempts.")

    return [{"描述": desc, "正则表达式": regex} for desc, regex in results]


def save_to_excel(dataset, filename="output.xlsx"):
    """保存数据集到 Excel 文件，并确保去重"""
    wb = Workbook()
    ws = wb.active

    # 写入表头
    ws["A1"] = "描述"
    ws["B1"] = "正则表达式"

    # 使用集合去重
    unique_entries = set()
    for entry in dataset:
        unique_key = (entry["描述"], entry["正则表达式"])
        unique_entries.add(unique_key)

    # 将去重后的数据写入 Excel
    for idx, (desc, regex) in enumerate(unique_entries, start=2):
        ws[f"A{idx}"] = desc
        ws[f"B{idx}"] = regex

    # 保存为 Excel 文件
    wb.save(filename)
    print(f"Dataset saved to {filename} with {len(unique_entries)} unique entries.")


if __name__ == "__main__":
    dataset = generate_dataset(2700)
    print(f"Generated {len(dataset)} unique entries.")
    save_to_excel(dataset, "./test_dataB.xlsx")