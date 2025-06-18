import openpyxl
import random

def read_excel(file_path):
    """读取Excel文件，返回数据列表（跳过标题行）"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if rows and rows[0][0] == "描述" and rows[0][1] == "正则表达式":
        rows = rows[1:]  # 跳过标题行
    return rows

def write_excel(data, file_path):
    """写入数据到Excel文件"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["描述", "正则表达式"])
    for row in data:
        sheet.append(row)
    workbook.save(file_path)

def merge_and_shuffle_excel(file_a, file_b, output_file):
    """读取两个Excel文件，合并并打乱数据，写入新文件"""
    print("正在读取 test_dataA.xlsx...")
    data_a = read_excel(file_a)
    print(f"读取到 {len(data_a)} 条数据")

    print("正在读取 test_dataB.xlsx...")
    data_b = read_excel(file_b)
    print(f"读取到 {len(data_b)} 条数据")

    combined = data_a + data_b
    print(f"总共数据条数: {len(combined)}")

    print("正在随机打乱数据...")
    random.shuffle(combined)

    print(f"正在写入文件: {output_file}")
    write_excel(combined, output_file)
    print("✅ 数据合并和打乱完成！")

if __name__ == "__main__":
    merge_and_shuffle_excel(
        file_a="test_dataA.xlsx",
        file_b="test_dataB.xlsx",
        output_file="test_data.xlsx"
    )