from modelscope import AutoModelForCausalLM, AutoTokenizer
import torch
from openpyxl import load_workbook
import time

# 当GPU0和GPU1空闲时可以使用此代码，用于多卡。单卡GPU0使用chat-25B_single_GPU.py

# 设备配置
device = "cuda" if torch.cuda.is_available() else "cpu"


# model_path = "Qwen/Qwen2.5-14B-Instruct"
# 指定本地缓存路径（根据实际路径修改）
model_cache_path = "/root/.cache/modelscope/hub/Qwen/Qwen2.5-14B-Instruct"

# 输入和输出文件路径
input_data = './data/regex_input.xlsx'
output_data = './data/regex_output.xlsx'

# 检查 GPU 数量
num_gpus = torch.cuda.device_count()
print(f"Detected {num_gpus} GPUs.")

# 加载模型和分词器，支持多 GPU 分布式加载
model = AutoModelForCausalLM.from_pretrained(
    model_cache_path,
    torch_dtype=torch.float16,  # 使用半精度浮点数以节省显存
    device_map="auto"         # 自动分配模型到多个 GPU
)
tokenizer = AutoTokenizer.from_pretrained(
    model_cache_path,
    padding_side='left'  # 设置左填充
    )

# 预热对话（固定上下文）
initial_messages = [
    {"role": "system", "content": "你是千问，由阿里云创造，你是一个有用的助手"},
    {"role": "user", "content": "给我一个关于大语言模型的简短介绍."},
    {"role": "assistant", "content": ""},
    {"role": "user", "content": '''你将根据已知的自然语言和它所描述的正则表达式来将复杂问题分解为子问题，给你一个示例如下：
    我输入：匹配电子邮件地址，它对应的正则表达式是[a-zA-Z0-9.-]+@[a-zA-Z0-9]+(-[a-zA-Z0-9]+)*
    你输出子问题步骤链：
    step1：匹配用户名部分，对应正则表达式：[a-zA-Z0-9._-]+
    step2：匹配 @ 符号，对应正则表达式：@
    step3：匹配域名部分，对应正则表达式：[a-zA-Z0-9]+(?:-[a-zA-Z0-9]+)*
    step4：组合以上步骤，对应正则表达式：[a-zA-Z0-9.-]+@[a-zA-Z0-9]+(-[a-zA-Z0-9]+)*
    你应该严格按照这样的格式来输出并且不需要使用?: 并且你只需要给我展示子问题步骤链，而不需要输出任何其它内容'''},
    {"role": "assistant", "content": ""},
    {"role": "user", "content": "按照我的格式，即step1，step2这样来展示思维链，展示最终的子问题步骤链即可"},
    {"role": "assistant", "content": ""},
    {"role": "user", "content": "并且最后一步必须是“组合以上步骤”"},
    {"role": "assistant", "content": ""}
]

def batch_talk(batch_inputs, line_numbers):
    """
    批量调用模型生成响应
    :param batch_inputs: 输入文本列表
    :param line_numbers: 当前处理的数据序号列表
    :return: 响应文本列表
    """
    start_time = time.time()

    # 显示当前处理的数据序号范围
    print(f"Processing lines {line_numbers[0]} to {line_numbers[-1]}")

    torch.cuda.empty_cache()  # 清理缓存
    cache_clear_time = time.time()

    # 构建完整的对话消息
    messages_list = [
        initial_messages + [{"role": "user", "content": input_text}]
        for input_text in batch_inputs
    ]
    build_messages_time = time.time()

    # 构建输入文本
    texts = [
        tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True,
        )
        for messages in messages_list
    ]
    apply_chat_template_time = time.time()

    # 将输入转换为模型可接受的格式
    model_inputs = tokenizer(texts, return_tensors="pt", padding=True, truncation=True).to(device)
    tokenization_time = time.time()

    # 生成输出
    generated_ids = model.generate(
        **model_inputs,
        max_new_tokens=512,
        temperature=0.5,     # 调整温度
        top_k=50,             # 设置 top-k 抽样
        top_p=0.5,           # 设置 nucleus 抽样
        use_cache=True      # 禁用缓存以避免显存泄漏
    )
    generation_time = time.time()

    # 解码生成的文本
    responses = tokenizer.batch_decode(
        [output_ids[len(input_ids):] for input_ids, output_ids in zip(model_inputs.input_ids, generated_ids)],
        skip_special_tokens=True
    )
    decoding_time = time.time()

    elapsed_time = time.time() - start_time
    print(f"Lines {line_numbers[0]} to {line_numbers[-1]} processed in {elapsed_time:.2f} seconds.")
    print(f"Cache cleared in {(cache_clear_time - start_time):.2f} seconds.")
    print(f"Messages built in {(build_messages_time - cache_clear_time):.2f} seconds.")
    print(f"Chat templates applied in {(apply_chat_template_time - build_messages_time):.2f} seconds.")
    print(f"Tokenized in {(tokenization_time - apply_chat_template_time):.2f} seconds.")
    print(f"Generated in {(generation_time - tokenization_time):.2f} seconds.")
    print(f"Decoded in {(decoding_time - generation_time):.2f} seconds.")
    
    return [resp.strip() for resp in responses]

def main():
    # 加载工作簿
    workbook = load_workbook(filename=input_data)
    sheet = workbook['Sheet']

    # 打印工作表名称
    print(workbook.sheetnames)

    # 设置起始行号（默认从第 1 行开始）
    line = 0
    batch_size = 10  # 增加批量大小以提高效率
    responses = []

    # 收集一批输入数据
    batch_inputs = []
    batch_line_numbers = []  # 存储当前批次的行号
    for row in sheet.iter_rows(min_row=line + 2, max_col=2, values_only=True):  # 跳过标题行
        line += 1
        description, regex = row

        # 如果目标列已经有内容，则跳过
        if sheet[f'C{line + 1}'].value:
            print(f"Skipping line {line} (already processed)")
            continue

        # 如果有空行则跳过
        if not description or not regex:
            print(f"Skipping empty row at line {line}")
            continue

        # 构造输入文本
        input_text = f"{description}，对应正则表达式：{regex}"
        batch_inputs.append(input_text)
        batch_line_numbers.append(line)

        # 如果达到批量大小，批量处理
        if len(batch_inputs) >= batch_size:
            batch_responses = batch_talk(batch_inputs, batch_line_numbers)  # 调用批量推理
            responses.extend(zip(batch_line_numbers, batch_responses))
            batch_inputs.clear()
            batch_line_numbers.clear()
            torch.cuda.empty_cache()

    # 处理剩余的数据
    if batch_inputs:
        batch_responses = batch_talk(batch_inputs, batch_line_numbers)  # 调用批量推理
        responses.extend(zip(batch_line_numbers, batch_responses))

    # 将结果写入 Excel 文件
    for l, resp in responses:
        sheet[f'C{l + 1}'] = resp
    workbook.save(filename=output_data)

    print("Success!")

if __name__ == "__main__":
    main()