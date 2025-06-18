from modelscope import AutoModelForCausalLM, AutoTokenizer
import torch
from openpyxl import load_workbook
import time
import psutil
import pandas as pd

# 设备配置
device = "cuda" if torch.cuda.is_available() else "cpu"

# 模型缓存路径
model_cache_path = "/root/.cache/modelscope/hub/qwen/Qwen2.5-7B-Instruct"

# 输入和输出文件路径
input_data = './data/test_data.xlsx'
output_data = './data/test_data_output.xlsx'
resource_usage_file = './data/resource_usage.xlsx'  # 资源使用情况输出路径

# 加载模型和分词器
model = AutoModelForCausalLM.from_pretrained(
    model_cache_path,
    torch_dtype=torch.float16,  # 使用半精度浮点数以节省显存
    device_map="auto"           # 自动分配模型到多个 GPU
)
tokenizer = AutoTokenizer.from_pretrained(model_cache_path)

# 预热对话（固定上下文）
initial_messages = [
    {"role": "system", "content": "你是千问，由阿里云创造，你是一个有用的助手"},
    {"role": "user", "content": "匹配电子邮件地址"},
    {"role": "assistant", "content": "[a-zA-Z0-9.-]+@[a-zA-Z0-9]+(-[a-zA-Z0-9]+)*"},
    {"role": "user", "content": "匹配2位数字"},
    {"role": "assistant", "content": "\\d{2}"},
    {"role": "user", "content": "我输入自然语言描述，请你直接输出对应的正则表达式，不要任何解释或额外内容。"},
    {"role": "assistant", "content": ""}
]

def generate_regex(input_text):
    """
    根据输入文本生成正则表达式
    :param input_text: 输入文本
    :return: 生成的正则表达式
    """
    # 构建完整的对话消息
    messages = initial_messages + [{"role": "user", "content": input_text}]
    
    # 构建输入文本
    text = tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True,
    )
    
    # 将输入转换为模型可接受的格式
    model_inputs = tokenizer(text, return_tensors="pt", padding=True, truncation=True).to(device)
    
    # 生成输出
    generated_ids = model.generate(
        **model_inputs,
        max_new_tokens=50,      # 限制生成长度，避免过长输出
        temperature=0.1,        # 降低温度以提高确定性
        top_k=10,               # 设置 top-k 抽样
        top_p=0.9,              # 设置 nucleus 抽样
        use_cache=True          # 使用缓存以提高效率
    )
    
    # 解码生成的文本
    response = tokenizer.decode(
        generated_ids[0][len(model_inputs.input_ids[0]):],  # 只解码生成的部分
        skip_special_tokens=True
    ).strip()
    
    return response

def contains_chinese(text):
    """
    检查字符串是否包含中文字符
    :param text: 输入字符串
    :return: 是否包含中文字符
    """
    for char in text:
        if '\u4e00' <= char <= '\u9fff':
            return True
    return False

def get_resource_usage():
    """
    获取当前进程的资源使用情况
    :return: 包含 CPU 使用率、内存 RSS 和 VMS 的字典
    """
    process = psutil.Process()
    cpu_percent = process.cpu_percent(interval=0.1) / psutil.cpu_count()
    memory_info = process.memory_info()
    return {
        'cpu_percent': cpu_percent,
        'memory_rss': memory_info.rss / (1024 * 1024),  # MB
        'memory_vms': memory_info.vms / (1024 * 1024)   # MB
    }

def main():
    # 加载 Excel 文件
    workbook = load_workbook(filename=input_data)
    sheet = workbook.active  # 获取活动工作表

    # 初始化资源数据列表
    resource_data = []

    # 遍历每一行数据（跳过标题行）
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # 从第2行开始
        description, correct_regex = row

        # 如果目标列已经有内容，则跳过
        if sheet[f'C{row_idx}'].value:
            print(f"Skipping line {row_idx} (already processed)")
            continue

        # 如果有空行则跳过
        if not description or not correct_regex:
            print(f"Skipping empty row at line {row_idx}")
            continue

        # 开始计时和记录资源
        start_time = time.time()
        resource_start = get_resource_usage()

        # 调用模型生成正则表达式
        generated_regex = generate_regex(description)

        # 结束时间和资源
        end_time = time.time()
        resource_end = get_resource_usage()

        elapsed_time = end_time - start_time
        avg_cpu = (resource_start['cpu_percent'] + resource_end['cpu_percent']) / 2
        avg_memory_rss = (resource_start['memory_rss'] + resource_end['memory_rss']) / 2
        avg_memory_vms = (resource_start['memory_vms'] + resource_end['memory_vms']) / 2

        # 记录资源使用情况
        resource_data.append({
            '描述': description,
            '处理时间': elapsed_time,
            'CPU使用率': avg_cpu,
            '内存使用RSS': avg_memory_rss,
            '内存使用VMS': avg_memory_vms
        })

        # 检查输出并写入 Excel
        if not generated_regex:
            print(f"Line {row_idx}: Empty output detected. Input: {description}")
        elif contains_chinese(generated_regex):
            print(f"Line {row_idx}: Chinese characters detected in output. Input: {description}, Output: {generated_regex}")

        print(f"Line {row_idx}: Input: {description}, Generated Regex: {generated_regex}")
        sheet[f'C{row_idx}'] = generated_regex

    # 保存修改后的 Excel 文件
    workbook.save(filename=output_data)
    print("Processing completed and saved to:", output_data)

    # 保存资源使用情况到新的 Excel 文件
    resource_df = pd.DataFrame(resource_data)
    resource_df.to_excel(resource_usage_file, index=False)
    print(f"资源使用情况已保存到 {resource_usage_file}")

if __name__ == "__main__":
    main()